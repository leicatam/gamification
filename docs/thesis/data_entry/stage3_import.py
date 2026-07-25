#!/usr/bin/env python3
"""stage3_import.py — load game state-table exports into the Stage-3 entry template.

Fills the RAW sheets of Stage3_Data_Entry_Template.xlsx from the game's export
files, so no value is typed cell by cell, and computes the exact stint-based
outcome flags beside the template's formula approximations.

Usage:
  python stage3_import.py --template Stage3_Data_Entry_Template.xlsx \
      [--participants participants.csv] [--sessions sessions.csv|.json] \
      [--states states.csv|.json] [--questionnaire q4.csv] \
      [--out Stage3_Data_Entry_FILLED.xlsx]

  python stage3_import.py --demo          # end-to-end pipeline test using the
                                          # SYNTHETIC workbook (TEST- IDs only)

Accepted inputs
  CSV with a header row, or JSON as a list of objects / {"sessions": [...]}.
  Column names are matched case-insensitively after stripping spaces and
  underscores, so exports like "participant id" or "PARTICIPANT_ID" both work.

Provenance rule
  This script moves data; it never creates it. Feed it only files exported
  from the game (Layer-3 logs) or transcribed from primary records.
"""
import argparse, csv, json, sys, datetime, re
from pathlib import Path

import openpyxl

def norm(s):
    return re.sub(r"[\s_]+", "", str(s).strip().lower())

def load_rows(path):
    """Return list[dict] from a CSV or JSON export."""
    p = Path(path)
    if p.suffix.lower() == ".json":
        data = json.loads(p.read_text())
        if isinstance(data, dict):
            for key in ("sessions", "states", "rows", "data", "records"):
                if key in data and isinstance(data[key], list):
                    data = data[key]; break
        if not isinstance(data, list):
            sys.exit(f"{path}: JSON must be a list of objects")
        return [dict(r) for r in data]
    with open(p, newline="", encoding="utf-8-sig") as f:
        return [dict(r) for r in csv.DictReader(f)]

# Known machine-export header synonyms (seen in the author's format sample
# telemetry_export_machine_v1.csv, 25 Jul 2026), mapped onto template columns.
ALIASES = {
    "machinetimestamp": "sessiondate",
    "hits": "obstacleshit",
    "difficulty": "aidiffmult",
    "gameover": "reachedgameover",
    "ageband": "ageband",
}

def remap(rows, wanted):
    """Map arbitrary export headers onto template columns (case/space-insensitive)."""
    out = []
    for r in rows:
        lut = {}
        for k, v in r.items():
            nk = norm(k)
            lut[ALIASES.get(nk, nk)] = v
        out.append([lut.get(norm(w), "") for w in wanted])
    return out

def parse_date(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    s = str(v).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try: return datetime.datetime.strptime(s, fmt).date()
        except ValueError: pass
    return None

PCOLS = ["Participant_ID","Recruit_Date","Age","Age_Band","Gender","Education",
         "Recruit_Channel","Prior_FRA_Use","Pre_FRA_Index","Pre_FES_I",
         "Consent_Signed","Withdrawn","Withdrawn_Reason"]
SCOLS = ["Participant_ID","Session_No","Session_Date","Duration_Min","Distance_m","Coins",
         "Obstacles_Hit","Lives_Lost","Avg_Balance_pct","Max_Speed","AI_DiffMult",
         "Reached_GameOver","Post_FRA_Index","Notes"]
GCOLS = ["Participant_ID","Session_No","Play_Time_s","Old_DiffMult","New_DiffMult","Reason"]
OCOLS = ["Participant_ID","Staff_Dropout_Observed","Dropout_Date",
         "Staff_Return_Observed","Return_Date","Observer_Initials","Notes"]
QCOLS = ["Participant_ID","Q4_Date","Q1_AI_Difficulty","Q2_Persistence","Q3_Features",
         "Q4_Enjoyment","Q5_SelfEfficacy","Q6_LearnReframe","Q7_CoordFrame",
         "Q8_ReturnRecommend","Q9_LikedMost","Q9_WouldChange"]

def coerce(v):
    """Give Excel real types: int, float, ISO date, boolean; else the string."""
    if v is None or isinstance(v, (int, float, datetime.date, datetime.datetime, bool)):
        return v
    s = str(v).strip()
    if s == "": return None
    if s.upper() == "TRUE": return True
    if s.upper() == "FALSE": return False
    if re.fullmatch(r"-?\d+", s): return int(s)
    if re.fullmatch(r"-?\d*\.\d+", s): return float(s)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return datetime.datetime.strptime(s, "%Y-%m-%d").date()
    return s

def write_sheet(ws, rows, start=2, skip_cols=()):
    for i, row in enumerate(rows):
        for j, v in enumerate(row, 1):
            if j in skip_cols: continue
            ws.cell(row=start+i, column=j, value=coerce(v))

def compute_stints(sessions, gap_days, five):
    """Exact stint logic: sessions sorted by date per participant; a gap
    > gap_days starts a new stint. Returns {pid: (first_stint_n, dropout, returned)}."""
    by = {}
    for s in sessions:
        pid, d = s[0], parse_date(s[2])
        if pid in ("", None) or d is None: continue
        by.setdefault(pid, []).append(d)
    out = {}
    for pid, dates in by.items():
        dates.sort()
        stints, cur = [], [dates[0]]
        for a, b in zip(dates, dates[1:]):
            if (b - a).days > gap_days: stints.append(cur); cur = [b]
            else: cur.append(b)
        stints.append(cur)
        first_n = len(stints[0])
        dropout = first_n < five
        returned = dropout and len(stints) > 1
        out[pid] = (first_n, dropout, returned)
    return out

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", default="Stage3_Data_Entry_Template.xlsx")
    ap.add_argument("--participants"); ap.add_argument("--sessions")
    ap.add_argument("--states"); ap.add_argument("--questionnaire")
    ap.add_argument("--observations", help="staff_observations.csv (dropout/return events)")
    ap.add_argument("--out", default=None)
    ap.add_argument("--demo", action="store_true",
                    help="pipeline test using the synthetic TEST- workbook")
    a = ap.parse_args()

    import openpyxl.styles.fonts as _f
    _f.Font.family.max = 200
    wb = openpyxl.load_workbook(a.template)

    if a.demo:
        src = openpyxl.load_workbook(
            "/root/.claude/uploads/3d8a50e8-9258-51ce-a7d6-194da50f9a04/"
            "fc547454-TestData_Stage3_Pipeline_v1_dates_corrected.xlsx")
        def sheet_rows(name, ncols, header_row=2):
            ws = src[name]
            rows = []
            for r in ws.iter_rows(min_row=header_row+1, max_col=ncols, values_only=True):
                if r[0] in (None, ""): continue
                rows.append(["" if v is None else v for v in r])
            return rows
        parts = sheet_rows("Participants", 13)
        sess  = sheet_rows("Game Telemetry", 14)
        states= sheet_rows("State_Changes", 6)
        q4    = sheet_rows("Q4 Responses", 12)
        outs  = sheet_rows("Outcomes", 10)
        obsrows = []
        # guard: demo data must be TEST- prefixed
        assert all(str(r[0]).startswith("TEST-") for r in parts), "demo expects TEST- IDs"
        print("DEMO MODE: synthetic TEST- data — pipeline verification only.")
    else:
        parts  = remap(load_rows(a.participants), PCOLS) if a.participants else []
        sess   = remap(load_rows(a.sessions), SCOLS) if a.sessions else []
        states = remap(load_rows(a.states), GCOLS) if a.states else []
        q4     = remap(load_rows(a.questionnaire), QCOLS) if a.questionnaire else []
        obsrows = remap(load_rows(a.observations), OCOLS) if a.observations else []

    if parts:
        ep = wb["ENTRY_Participants"]
        # column D is the Age_Band formula — do not overwrite (skip col 4);
        # participant IDs in col A are overwritten by the import's own IDs.
        write_sheet(ep, parts, skip_cols=(4,))
        print(f"participants written: {len(parts)}")
    if sess:
        write_sheet(wb["RAW_Sessions"], sess)
        print(f"sessions written: {len(sess)}")
    if states:
        write_sheet(wb["RAW_GameStates"], states)
        print(f"state changes written: {len(states)}")
    if q4:
        write_sheet(wb["RAW_Questionnaire"], q4)
        print(f"questionnaires written: {len(q4)}")
    if not a.demo and obsrows:
        eo = wb["ENTRY_Observations"]
        # align observation rows to participant order; unknown IDs appended below
        order = {p[0]: i for i, p in enumerate(parts)}
        extra = len(parts)
        for orow in obsrows:
            i = order.get(orow[0])
            if i is None:
                i = extra; extra += 1
            row = 2 + i
            eo.cell(row=row, column=1, value=orow[0])
            for j, v in enumerate(orow[1:], start=2):
                b = str(v).strip().upper()
                eo.cell(row=row, column=j,
                        value=(True if b == "TRUE" else False if b == "FALSE" else (v or None)))
        print(f"staff observations written: {len(obsrows)}")
    if a.demo and outs:
        eo = wb["ENTRY_Observations"]
        idx = {r[0]: r for r in outs}
        for i, prow in enumerate(parts):
            rec = idx.get(prow[0])
            if rec is None: continue
            row = 2 + i
            eo.cell(row=row, column=1, value=prow[0])
            eo.cell(row=row, column=2, value=str(rec[2]).upper() == "TRUE")   # Initial_Dropout
            eo.cell(row=row, column=4, value=str(rec[3]).upper() == "TRUE")   # Voluntary_Return
            eo.cell(row=row, column=6, value="SIM")
        print(f"staff observations (simulated from synthetic Outcomes): {len(outs)}")

    # exact stint-based outcomes -> AUTO_Outcomes columns L,M,N
    cfg = {r[0].value: r[1].value for r in wb["Config"].iter_rows(min_row=2)}
    gap = int(cfg.get("RETURN_GAP_DAYS", 2)); five = int(cfg.get("FIVE_ATTEMPTS", 5))
    stints = compute_stints(sess, gap, five)
    ao = wb["AUTO_Outcomes"]
    pids = [r[0] for r in parts]
    for i, pid in enumerate(pids):
        row = 2 + i
        ao.cell(row=row, column=1, value=pid)   # replace formula-linked ID with actual
        fs, dr, rt = stints.get(pid, (0, False, False))
        ao.cell(row=row, column=15, value=fs)   # Script_First_Stint
        ao.cell(row=row, column=16, value=rt)   # Script_Stint_Return

    n_drop = sum(1 for v in stints.values() if v[1])
    n_ret  = sum(1 for v in stints.values() if v[2])
    print(f"\nScript-computed outcomes: participants with sessions={len(stints)}, "
          f"initial dropouts={n_drop}, voluntary returns={n_ret}, "
          f"return rate={n_ret/max(n_drop,1)*100:.1f}%")

    out = a.out or ("Stage3_Data_Entry_DEMO_SYNTHETIC.xlsx" if a.demo
                    else "Stage3_Data_Entry_FILLED.xlsx")
    wb.save(out)
    print("saved:", out)

if __name__ == "__main__":
    main()
