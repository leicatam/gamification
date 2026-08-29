#!/usr/bin/env python3
"""Rebuild the Stage-3 master logbook around the ACTUAL record trail of
the remote deployment (author's account, 25 Jul 2026): each participant
e-mailed back the HTML build's logged results CSV and their
questionnaire responses. Adds an Email_Returns_Log entry sheet and a
RAW_Email_CSV paste sheet to the existing master logbook, and prepends a
provenance note to the Codebook. Existing sheets are left in place."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

P = "/home/user/gamification/docs/thesis/Stage3_Data_Logbook_120cases_with_Reconciliation.xlsx"
wb = openpyxl.load_workbook(P)

RED = Font(color="9C0006", bold=True)
HDR = Font(bold=True)
FILL = PatternFill("solid", fgColor="FCE4E4")
NOTE = ("PROVENANCE (redefined 25 Jul 2026): Stage 3 was a REMOTE "
        "deployment — an HTML build of the game e-mailed to participants, "
        "played on their own computers (trackpad / arrow keys / foot-pad "
        "sensor where available). The primary records are the "
        "participants' RETURNED E-MAILS: the build's logged results CSV "
        "(incl. score) and the questionnaire responses. Fill this logbook "
        "ONLY by transcription from those returned e-mails. No terminal "
        "exports, paper questionnaires or staff session logs exist for "
        "Stage 3. Every row must be traceable to a specific returned "
        "e-mail (date + filename in Email_Returns_Log).")

# 1) Codebook provenance note
cb = wb["Codebook"]
cb.insert_rows(1, 2)
c = cb.cell(row=1, column=1, value=NOTE)
c.font = RED
c.alignment = Alignment(wrap_text=True, vertical="top")
cb.row_dimensions[1].height = 90

# 2) Email_Returns_Log
if "Email_Returns_Log" in wb.sheetnames:
    del wb["Email_Returns_Log"]
ws = wb.create_sheet("Email_Returns_Log", 1)
ws.cell(row=1, column=1, value=(
    "ENTRY POINT for the Stage-3 compilation. One row per participant. "
    "Transcribe from the actual returned e-mails only; leave blank where "
    "no return exists (blanks are findings, not gaps to fill)."))
ws.cell(row=1, column=1).font = RED
ws.cell(row=1, column=1).fill = FILL
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
ws.row_dimensions[1].height = 40
ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
headers = ["Participant_ID", "Distribution_Channel (Work/Charity/University/MedicalCentre)",
           "Build_Sent_Date", "Return_Email_Date", "CSV_Filename",
           "Questionnaire_Returned (Y/N)", "Input_Device (Trackpad/ArrowKeys/FootPad)",
           "Score_Reported", "Notes"]
for j, h in enumerate(headers, 1):
    ws.cell(row=2, column=j, value=h).font = HDR
for i in range(120):
    ws.cell(row=3 + i, column=1, value=3001 + i)
ws.column_dimensions["A"].width = 14
for col in "BCDEFGHI":
    ws.column_dimensions[col].width = 24

# 3) RAW_Email_CSV paste sheet
if "RAW_Email_CSV" in wb.sheetnames:
    del wb["RAW_Email_CSV"]
ws2 = wb.create_sheet("RAW_Email_CSV", 2)
ws2.cell(row=1, column=1, value=(
    "Paste the rows of each participant's returned CSV here UNCHANGED, "
    "prefixed with the Participant_ID and the source filename. Do not "
    "edit, reorder or 'clean' values. The header row of the build's CSV "
    "goes in row 2 once the first genuine return is pasted."))
ws2.cell(row=1, column=1).font = RED
ws2.cell(row=1, column=1).fill = FILL
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
ws2.row_dimensions[1].height = 55
ws2.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
ws2.cell(row=2, column=1, value="Participant_ID").font = HDR
ws2.cell(row=2, column=2, value="Source_CSV_Filename").font = HDR
ws2.cell(row=2, column=3, value="<build CSV columns from first genuine return>").font = HDR

wb.save(P)
print("Logbook rebuilt:", wb.sheetnames)
