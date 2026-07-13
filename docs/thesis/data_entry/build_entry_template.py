#!/usr/bin/env python3
"""Build the Stage-3 data-entry template workbook.

Design:
  HowTo             - instructions and definitions
  Config            - editable thresholds & rules (definitions live HERE, not in formulas)
  ENTRY_Participants- typed enrolment entry with dropdown validation, Age_Band auto
  RAW_Sessions      - paste target matching the game's per-session export
  RAW_GameStates    - paste target matching the game's state-change (L2 difficulty) export
  RAW_Questionnaire - paste target for the post-study questionnaire export
  AUTO_Outcomes     - one row per participant, every cell derived by formula
  CHECKS            - live data-quality flags (duplicates, orphans, out-of-range)

Nothing in this workbook invents data: AUTO sheets only derive from what is
pasted into RAW sheets; CHECKS only flags problems.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "Stage3_Data_Entry_Template.xlsx"
N = 120          # participant capacity
SMAX = 1200      # session rows capacity (allows re-visits well beyond 5)

wb = openpyxl.Workbook()
hdr = Font(bold=True)
hfill = PatternFill("solid", fgColor="DDEBF7")
afill = PatternFill("solid", fgColor="E2EFDA")   # auto-derived
warn = PatternFill("solid", fgColor="FCE4EC")

# ------------------------------------------------------------------ HowTo
ht = wb.active
ht.title = "HowTo"
lines = [
    ("Stage-3 Data-Entry Template — how to use", True),
    ("", False),
    ("1. ENTER participants once, in ENTRY_Participants. Age_Band fills itself; dropdowns constrain categories.", False),
    ("2. PASTE the game's per-session export into RAW_Sessions (columns already match the game state tables).", False),
    ("   If you have raw JSON logs instead of a spreadsheet export, run stage3_import.py — it fills the RAW sheets for you.", False),
    ("3. PASTE the L2 difficulty state-change export into RAW_GameStates (optional but recommended: it evidences the adaptive layer).", False),
    ("4. PASTE questionnaire responses into RAW_Questionnaire.", False),
    ("5. AUTO_Outcomes computes itself. Do not type into green cells - they are formulas.", False),
    ("6. Watch CHECKS: every flag must read 0 before the data moves to the master logbook.", False),
    ("7. When CHECKS is clean, copy AUTO_Outcomes (paste values) into the master logbook's Outcomes sheet and the RAW sheets into its data sheets. The logbook's Reconciliation sheet then compares your totals with the published Paper 3 aggregates.", False),
    ("", False),
    ("DEFINITIONS are set on the Config sheet and are used by every formula. If a definition must change, change it there once - never by editing formulas - and record the change in the thesis (§6.2/§8.9).", True),
    ("", False),
    ("Definition notes:", True),
    ("- Initial_Dropout (formula) = participant has at least one session but fewer than FIVE_ATTEMPTS sessions in their first stint.", False),
    ("  The formula approximates the first stint by counting sessions dated within RETURN_GAP_DAYS of the previous session;", False),
    ("  stage3_import.py computes the exact stint-based value and writes it beside the formula for comparison.", False),
    ("- Voluntary_Return = any session that starts a new stint (gap > RETURN_GAP_DAYS) after an incomplete first stint.", False),
    ("- Coord_Endorse / Regular_Use_Intent = questionnaire item >= threshold on Config.", False),
    ("  NOTE: thesis Appendix D specifies a 5-point scale with endorsement = 4 or 5; the logbook codebook says Likert 1-7.", False),
    ("  Config defaults to the 7-point reading (>=5). Confirm which scale the questionnaire actually used and set Config to match", False),
    ("  BEFORE entering data; record the decision in Appendix D.", False),
    ("- Dropout/return PRECEDENCE: staff-observed events (ENTRY_Observations, transcribed from staff session logs) override", False),
    ("  telemetry signals, because a participant can drop out, return, and still finish all five sessions - telemetry alone", False),
    ("  cannot see that. FINAL_* columns apply this precedence; CHECKS flags disagreements for investigation.", False),
    ("- Data provenance rule: RAW sheets may contain only values transcribed or imported from primary records/game logs.", True),
]
for i, (t, b) in enumerate(lines, 1):
    c = ht.cell(row=i, column=1, value=t)
    c.font = Font(bold=b)
    c.alignment = Alignment(wrap_text=False)
ht.column_dimensions["A"].width = 120

# ------------------------------------------------------------------ Config
cf = wb.create_sheet("Config")
cf.append(["Setting", "Value", "Meaning"])
for c in cf[1]: c.font = hdr; c.fill = hfill
config_rows = [
    ("FIVE_ATTEMPTS", 5, "Number of attempts defining in-session completion (five-attempt rule)"),
    ("RETURN_GAP_DAYS", 10, "A gap of MORE than this many days starts a new stint; set it comfortably above the normal session cadence"),
    ("ENDORSE_ITEM", "Q7_CoordFrame", "Questionnaire item used for coordination endorsement"),
    ("ENDORSE_THRESHOLD", 5, "Minimum response counting as endorsement (7-point scale default; use 4 if 5-point)"),
    ("INTENT_ITEM", "Q8_ReturnRecommend", "Questionnaire item used for regular-use intention"),
    ("INTENT_THRESHOLD", 5, "Minimum response counting as intention (7-point default; use 4 if 5-point)"),
    ("STUDY_START", "2026-03-01", "First valid session date (sessions before this are flagged)"),
    ("STUDY_END", "2026-05-31", "Last valid session date (sessions after this are flagged)"),
]
for r in config_rows: cf.append(r)
cf.column_dimensions["A"].width = 22; cf.column_dimensions["B"].width = 16; cf.column_dimensions["C"].width = 90
# named cells via defined names
from openpyxl.workbook.defined_name import DefinedName
for i, (name, *_ ) in enumerate(config_rows, start=2):
    wb.defined_names[name] = DefinedName(name, attr_text=f"Config!$B${i}")

# ------------------------------------------------------------------ ENTRY_Participants
ep = wb.create_sheet("ENTRY_Participants")
pcols = ["Participant_ID","Recruit_Date","Age","Age_Band (auto)","Gender","Education",
         "Recruit_Channel","Prior_FRA_Use","Pre_FRA_Index","Pre_FES_I","Consent_Signed",
         "Withdrawn","Withdrawn_Reason"]
ep.append(pcols)
for c in ep[1]: c.font = hdr; c.fill = hfill
for r in range(2, N+2):
    ep.cell(row=r, column=1, value=3000+r-1)            # 3001..3120 pre-filled
    f = ep.cell(row=r, column=4, value=f'=IF($C{r}="","",IF($C{r}<=44,"25-44",IF($C{r}<=64,"45-64","65-80")))')
    f.fill = afill
dvs = {
    "E": '"M,F,Other,NA"',
    "F": '"Primary,Secondary,Tertiary,Postgrad"',
    "H": '"Yes,No,Unknown"',
    "K": '"TRUE,FALSE"',
    "L": '"TRUE,FALSE"',
}
for col, formula in dvs.items():
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=True)
    ep.add_data_validation(dv)
    dv.add(f"{col}2:{col}{N+1}")
for i, w in enumerate([14,12,6,12,8,11,15,13,13,10,14,10,18], 1):
    ep.column_dimensions[get_column_letter(i)].width = w

# ------------------------------------------------------------------ RAW_Sessions
rs = wb.create_sheet("RAW_Sessions")
scols = ["Participant_ID","Session_No","Session_Date","Duration_Min","Distance_m","Coins",
         "Obstacles_Hit","Lives_Lost","Avg_Balance_pct","Max_Speed","AI_DiffMult",
         "Reached_GameOver","Post_FRA_Index","Notes"]
rs.append(scols)
for c in rs[1]: c.font = hdr; c.fill = hfill
dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
rs.add_data_validation(dv); dv.add(f"L2:L{SMAX+1}")

# ------------------------------------------------------------------ RAW_GameStates
rg = wb.create_sheet("RAW_GameStates")
gcols = ["Participant_ID","Session_No","Play_Time_s","Old_DiffMult","New_DiffMult","Reason"]
rg.append(gcols)
for c in rg[1]: c.font = hdr; c.fill = hfill

# ------------------------------------------------------------------ RAW_Questionnaire
rq = wb.create_sheet("RAW_Questionnaire")
qcols = ["Participant_ID","Q4_Date","Q1_AI_Difficulty","Q2_Persistence","Q3_Features",
         "Q4_Enjoyment","Q5_SelfEfficacy","Q6_LearnReframe","Q7_CoordFrame",
         "Q8_ReturnRecommend","Q9_LikedMost","Q9_WouldChange"]
rq.append(qcols)
for c in rq[1]: c.font = hdr; c.fill = hfill
dvq = DataValidation(type="whole", operator="between", formula1="1", formula2="7",
                     allow_blank=True, showErrorMessage=True,
                     error="Likert items must be 1-7 (set Config if using a 5-point scale)")
rq.add_data_validation(dvq); dvq.add(f"C2:J{N+1}")


# ------------------------------------------------------------------ ENTRY_Observations
eo = wb.create_sheet("ENTRY_Observations")
obcols = ["Participant_ID","Staff_Dropout_Observed","Dropout_Date","Staff_Return_Observed",
          "Return_Date","Observer_Initials","Notes"]
eo.append(obcols)
for c in eo[1]: c.font = hdr; c.fill = hfill
for r in range(2, N+2):
    eo.cell(row=r, column=1, value=f"=ENTRY_Participants!$A{r}")
dvo = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
eo.add_data_validation(dvo); dvo.add(f"B2:B{N+1}"); dvo.add(f"D2:D{N+1}")
for i, w in enumerate([14,22,13,21,12,16,30], 1):
    eo.column_dimensions[get_column_letter(i)].width = w

# ------------------------------------------------------------------ AUTO_Outcomes
ao = wb.create_sheet("AUTO_Outcomes")
ocols = ["Participant_ID","Sessions_Total","Completed_Session1","First_Stint (formula approx)",
         "Telemetry_InSession_Retention","Telemetry_Dropout_Signal",
         "Staff_Dropout_Observed","Staff_Return_Observed",
         "FINAL_Initial_Dropout","FINAL_Voluntary_Return",
         "Coord_Endorse","Regular_Use_Intent","Last_Post_FRA_Index","Post_minus_Pre_FRA",
         "Script_First_Stint","Script_Stint_Return"]
ao.append(ocols)
for c in ao[1]: c.font = hdr; c.fill = hfill
SA = f"RAW_Sessions!$A$2:$A${SMAX+1}"
SB = f"RAW_Sessions!$B$2:$B${SMAX+1}"
SC = f"RAW_Sessions!$C$2:$C${SMAX+1}"
SM = f"RAW_Sessions!$M$2:$M${SMAX+1}"
QA = f"RAW_Questionnaire!$A$2:$A${N+1}"
for r in range(2, N+2):
    ID = f"ENTRY_Participants!$A{r}"
    ao.cell(row=r, column=1, value=f"={ID}")
    ao.cell(row=r, column=2, value=f"=COUNTIF({SA},{ID})")
    ao.cell(row=r, column=3, value=f"=IF(COUNTIFS({SA},{ID},{SB},1)>0,TRUE,FALSE)")
    ao.cell(row=r, column=4, value=(
        f"=IF(B{r}=0,0,COUNTIFS({SA},{ID},{SC},\"<=\"&MINIFS({SC},{SA},{ID})+RETURN_GAP_DAYS*FIVE_ATTEMPTS))"))
    ao.cell(row=r, column=5, value=f"=IF(B{r}=0,FALSE,COUNTIFS({SA},{ID},{SB},\"<=\"&FIVE_ATTEMPTS)>=FIVE_ATTEMPTS)")
    ao.cell(row=r, column=6, value=f"=IF(B{r}=0,FALSE,NOT(E{r}))")
    ao.cell(row=r, column=7, value=f"=IF(ENTRY_Observations!$B{r}=\"\",\"\",ENTRY_Observations!$B{r})")
    ao.cell(row=r, column=8, value=f"=IF(ENTRY_Observations!$D{r}=\"\",\"\",ENTRY_Observations!$D{r})")
    ao.cell(row=r, column=9, value=f"=IF(G{r}<>\"\",G{r},F{r})")
    ao.cell(row=r, column=10, value=f"=IF(H{r}<>\"\",H{r},IF(I{r}=FALSE,FALSE,P{r}))")
    ao.cell(row=r, column=11, value=(
        f"=IF(COUNTIF({QA},{ID})=0,\"\",IF(SUMIFS(RAW_Questionnaire!$I$2:$I${N+1},{QA},{ID})>=ENDORSE_THRESHOLD,TRUE,FALSE))"))
    ao.cell(row=r, column=12, value=(
        f"=IF(COUNTIF({QA},{ID})=0,\"\",IF(SUMIFS(RAW_Questionnaire!$J$2:$J${N+1},{QA},{ID})>=INTENT_THRESHOLD,TRUE,FALSE))"))
    ao.cell(row=r, column=13, value=f"=IFERROR(LOOKUP(2,1/(({SA}={ID})*({SM}<>\"\")),{SM}),\"\")")
    ao.cell(row=r, column=14, value=f"=IF(OR(M{r}=\"\",ENTRY_Participants!$I{r}=\"\"),\"\",M{r}-ENTRY_Participants!$I{r})")
    for col in range(2, 15):
        ao.cell(row=r, column=col).fill = afill
for i, w in enumerate([14,13,17,24,26,22,21,20,19,21,13,17,17,17,15,16], 1):
    ao.column_dimensions[get_column_letter(i)].width = w

# ------------------------------------------------------------------ CHECKS
ck = wb.create_sheet("CHECKS")
ck.append(["Check", "Count (must be 0)", "Where to look"])
for c in ck[1]: c.font = hdr; c.fill = hfill
PA = f"ENTRY_Participants!$A$2:$A${N+1}"
PC = f"ENTRY_Participants!$C$2:$C${N+1}"
checks = [
    ("Duplicate participant IDs", f"=SUMPRODUCT((COUNTIF({PA},{PA})>1)*1)", "ENTRY_Participants"),
    ("Ages outside 25-80", f"=COUNTIF({PC},\"<25\")+COUNTIF({PC},\">80\")", "ENTRY_Participants col C"),
    ("Sessions with ID not in participants", f"=SUMPRODUCT(({SA}<>\"\")*(COUNTIF({PA},{SA})=0))", "RAW_Sessions col A"),
    ("Duplicate participant+session pairs", f"=SUMPRODUCT(({SA}<>\"\")*(COUNTIFS({SA},{SA},{SB},{SB})>1))", "RAW_Sessions"),
    ("Session dates before STUDY_START", f"=COUNTIFS({SC},\"<\"&DATEVALUE(STUDY_START),{SC},\"<>\")", "RAW_Sessions col C"),
    ("Session dates after STUDY_END", f"=COUNTIF({SC},\">\"&DATEVALUE(STUDY_END))", "RAW_Sessions col C"),
    ("Obstacles_Hit outside 0-3", f"=COUNTIF(RAW_Sessions!$G$2:$G${SMAX+1},\">3\")+COUNTIF(RAW_Sessions!$G$2:$G${SMAX+1},\"<0\")", "RAW_Sessions col G"),
    ("Balance % outside 0-100", f"=COUNTIF(RAW_Sessions!$I$2:$I${SMAX+1},\">100\")+COUNTIF(RAW_Sessions!$I$2:$I${SMAX+1},\"<0\")", "RAW_Sessions col I"),
    ("Post FRA index outside 0-100", f"=COUNTIF({SM},\">100\")+COUNTIF({SM},\"<0\")", "RAW_Sessions col M"),
    ("Questionnaire IDs not in participants", f"=SUMPRODUCT(({QA}<>\"\")*(COUNTIF({PA},{QA})=0))", "RAW_Questionnaire col A"),
    ("Duplicate questionnaire IDs", f"=SUMPRODUCT(({QA}<>\"\")*(COUNTIF({QA},{QA})>1))", "RAW_Questionnaire"),
    ("Staff dropout record contradicts telemetry (staff FALSE but <5 sessions)", f"=SUMPRODUCT((AUTO_Outcomes!$G$2:$G${N+1}=FALSE)*(AUTO_Outcomes!$F$2:$F${N+1}=TRUE))", "AUTO_Outcomes G vs F"),
    ("Return recorded without dropout", f"=SUMPRODUCT((AUTO_Outcomes!$J$2:$J${N+1}=TRUE)*(AUTO_Outcomes!$I$2:$I${N+1}=FALSE))", "AUTO_Outcomes I vs J"),
    ("Participants with sessions but no session 1", f"=SUMPRODUCT((COUNTIF({SA},{PA})>0)*(COUNTIFS({SA},{PA},{SB},1)=0))", "cross-check"),
]
for name, f, where in checks:
    ck.append([name, f, where])
for r in range(2, len(checks)+2):
    ck.cell(row=r, column=2).fill = warn
ck.column_dimensions["A"].width = 44; ck.column_dimensions["B"].width = 18; ck.column_dimensions["C"].width = 30

wb.save(OUT)
print("saved", OUT)
